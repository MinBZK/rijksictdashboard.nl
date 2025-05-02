import io
from dataclasses import dataclass

import pandas as pd
from openpyxl import Workbook, styles
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from pyexcel_ods3 import save_data
from pyexcel_xlsx import get_data

from app.routers.dependencies.excel_config import ExcelCell
from app.types.misc import SpreadsheetType


def dict_to_df(data_dict: dict) -> dict[str, pd.DataFrame]:
    """
    Returns a dictionary with sheet name as key and sheet data as a DataFrame.

    The data dict can contain primitive values (string, float, etc.), list, or unnested dictionary.
    """

    primitive_items = {
        k: v
        for k, v in data_dict.items()
        if not isinstance(v, dict) and not isinstance(v, list)
    }
    non_primitive_items = {
        k: v for k, v in data_dict.items() if isinstance(v, dict) or isinstance(v, list)
    }
    df_algemeen = pd.DataFrame.from_dict(primitive_items, orient="index").reset_index()
    if len(df_algemeen.index) > 0:
        df_algemeen.columns = ["Rapportage item", "Waarde"]
        df_collection = {"Algemeen": df_algemeen}
    else:
        df_collection = {}

    for k, v in non_primitive_items.items():
        if isinstance(v, dict) and len(v.keys()) > 0:
            df_collection[k] = pd.DataFrame.from_dict(v, orient="index").reset_index()  # type: ignore
            df_collection[k].columns = ["Parameter", "Waarde"]
            df_collection[k].sort_values(by=["Parameter"], inplace=True)
        elif not isinstance(v, dict):
            df_collection[k] = pd.DataFrame(v)

    return df_collection


@dataclass
class ExcelSheet:
    title: str
    cells: list[ExcelCell]


def df_dict_to_excel_stream(
    df_dict: dict, sheet_names: dict = {}, default_sheet: ExcelSheet | None = None
) -> io.BytesIO:
    """
    Transforms a dictionary of dataframes into an Excel file, where
    each dataframe is on a seperate sheet.

    Returns the Excel file as a stream.
    """
    stream = io.BytesIO()
    wb = Workbook()
    column_width = 25

    for title, df in df_dict.items():
        if len(df.index) > 0:
            ws = wb.create_sheet(title=sheet_names.get(title, title))
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

            for col in ws.columns:
                ws.column_dimensions[
                    get_column_letter(col[0].column)
                ].width = column_width

            for cell in ws[1]:
                cell.style = "Pandas"

    if default_sheet is not None:
        ws = wb["Sheet"]
        ws.title = default_sheet.title

        for col in ws.columns:
            if col[0].column:
                ws.column_dimensions[
                    get_column_letter(col[0].column)
                ].width = column_width

        for index, cell in enumerate(default_sheet.cells):
            cell_label = f"A{index + 1}"
            ws[cell_label] = cell.body if cell.body is not None else ""
            ws[cell_label].alignment = styles.Alignment(wrap_text=True)
            if cell.bold:
                ws[cell_label].font = styles.Font(bold=True)
    else:
        del wb["Sheet"]

    wb.save(stream)

    return stream


def excel_to_ods(excel_file: io.BytesIO) -> io.BytesIO:
    bytes_ods = io.BytesIO()
    save_data(bytes_ods, get_data(excel_file))
    return bytes_ods


def dict_to_spreadsheet(
    data_dict: dict,
    file_type: SpreadsheetType,
    sheet_names: dict = {},
    default_sheet: ExcelSheet | None = None,
) -> io.BytesIO:
    df_dict = dict_to_df(data_dict=data_dict)
    stream = df_dict_to_excel_stream(
        df_dict=df_dict, sheet_names=sheet_names, default_sheet=default_sheet
    )

    return stream if file_type == "excel" else excel_to_ods(stream)
